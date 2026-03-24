"""
Property Index Project Explorer — Server
=========================================
Auth + storage powered by Supabase.
Wallet system: tokens purchased via plans, deducted on project unlock.
"""

from flask import Flask, jsonify, session, request, send_from_directory, abort
import openpyxl, json, os, math
from functools import wraps
from datetime import datetime, timezone, timedelta
from config import EXCEL_PATH, GRAPHS_DIR, SECRET_KEY, PORT, SUPABASE_URL, SUPABASE_ANON_KEY, SUPABASE_SERVICE_KEY
from supabase import create_client, Client

app = Flask(__name__, static_folder="static")
app.secret_key = SECRET_KEY

supabase: Client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
# Admin client (service role) — used for profile updates & wallet writes
supabase_admin: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY) if SUPABASE_SERVICE_KEY else supabase

# ─── Pricing / token config ──────────────────────────────
PROJECT_TOKEN_COST = 100        # tokens to unlock one project

PLANS = {
    "starter": {
        "name":       "Starter Plan",
        "tokens":     500,
        "price_inr":  500,
        "unlock_days": 15,
    },
    "enterprise": {
        "name":       "Enterprise Plan",
        "tokens":     5000,
        "price_inr":  5000,
        "unlock_days": 30,
    },
}

# unlock duration by source: direct (no plan) = 5 days
DIRECT_UNLOCK_DAYS    = 5
STARTER_UNLOCK_DAYS   = 15
ENTERPRISE_UNLOCK_DAYS = 30

# ─── Auth helpers ────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            abort(401)
        return f(*args, **kwargs)
    return decorated

def clean(val):
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()

# ─── Wallet helpers (Supabase `wallets` table) ────────────
# Schema: wallets(user_id PK, tokens int, plan text, plan_activated_at timestamptz)

def _wallet_row(user_id: str) -> dict:
    """Return wallet row or default zero wallet."""
    res = supabase.table("wallets").select("*").eq("user_id", user_id).execute()
    if res.data:
        return res.data[0]
    return {"user_id": user_id, "tokens": 0, "plan": None, "plan_activated_at": None}

def get_wallet(user_id: str) -> dict:
    row = _wallet_row(user_id)
    plan = row.get("plan")
    plan_activated_at = row.get("plan_activated_at")
    plan_info = None
    if plan and plan_activated_at:
        activated = datetime.fromisoformat(plan_activated_at.replace("Z", "+00:00"))
        plan_cfg  = PLANS.get(plan, {})
        plan_info = {
            "plan":          plan,
            "name":          plan_cfg.get("name", plan),
            "activated_at":  plan_activated_at,
            "unlock_days":   plan_cfg.get("unlock_days", DIRECT_UNLOCK_DAYS),
        }
    return {
        "tokens":   row.get("tokens", 0),
        "plan":     plan_info,
    }

def deduct_tokens(user_id: str, amount: int) -> bool:
    """Deduct tokens; returns False if insufficient balance."""
    row = _wallet_row(user_id)
    current = row.get("tokens", 0)
    if current < amount:
        return False
    new_balance = current - amount
    supabase_admin.table("wallets").upsert({
        "user_id": user_id,
        "tokens":  new_balance,
        "plan":    row.get("plan"),
        "plan_activated_at": row.get("plan_activated_at"),
    }).execute()
    return True

def add_tokens(user_id: str, amount: int, plan: str | None = None):
    """Add tokens to wallet; optionally record the plan."""
    row    = _wallet_row(user_id)
    update = {
        "user_id": user_id,
        "tokens":  row.get("tokens", 0) + amount,
        "plan":    plan if plan else row.get("plan"),
        "plan_activated_at": datetime.now(timezone.utc).isoformat() if plan else row.get("plan_activated_at"),
    }
    supabase_admin.table("wallets").upsert(update).execute()

# ─── Unlocks (Supabase Postgres, expiry depends on unlock source) ─────
# Schema: unlocks(user_id, project_index, unlocked_at, unlock_source text)
# unlock_source: "direct" | "starter" | "enterprise"

def _unlock_days_for_source(source: str) -> int:
    if source == "starter":
        return STARTER_UNLOCK_DAYS
    if source == "enterprise":
        return ENTERPRISE_UNLOCK_DAYS
    return DIRECT_UNLOCK_DAYS   # "direct" or legacy

def _unlock_row(user_id: str, project_index: int):
    """Return unlock row if it exists and is not expired, else None."""
    res = supabase.table("unlocks") \
        .select("project_index, unlocked_at, unlock_source") \
        .eq("user_id", user_id) \
        .eq("project_index", project_index) \
        .execute()
    if not res.data:
        return None
    row     = res.data[0]
    source  = row.get("unlock_source", "direct")
    days    = _unlock_days_for_source(source)
    unlocked_at = datetime.fromisoformat(row["unlocked_at"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) - unlocked_at > timedelta(days=days):
        return None   # expired
    return row

def is_unlocked(user_id: str, project_index: int) -> bool:
    return _unlock_row(user_id, project_index) is not None

def unlock_project(user_id: str, project_index: int, source: str = "direct"):
    """Insert or refresh the unlock with source (resets expiry clock)."""
    supabase.table("unlocks").upsert({
        "user_id":        user_id,
        "project_index":  project_index,
        "unlocked_at":    datetime.now(timezone.utc).isoformat(),
        "unlock_source":  source,
    }).execute()

def get_unlocked_list(user_id: str) -> list:
    """Return list of {project_index, unlocked_at, days_left, unlock_source} for non-expired unlocks."""
    res = supabase.table("unlocks") \
        .select("project_index, unlocked_at, unlock_source") \
        .eq("user_id", user_id) \
        .execute()
    now    = datetime.now(timezone.utc)
    result = []
    for row in res.data:
        source      = row.get("unlock_source", "direct")
        days        = _unlock_days_for_source(source)
        unlocked_at = datetime.fromisoformat(row["unlocked_at"].replace("Z", "+00:00"))
        elapsed     = now - unlocked_at
        if elapsed <= timedelta(days=days):
            days_left = days - int(elapsed.total_seconds() // 86400)
            result.append({
                "project_index": row["project_index"],
                "unlocked_at":   row["unlocked_at"],
                "days_left":     days_left,
                "unlock_source": source,
            })
    return result

def get_unlocked_set(user_id: str) -> set:
    return {r["project_index"] for r in get_unlocked_list(user_id)}

# ─── Data loading ─────────────────────────────────────────

def load_projects():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    headers = [clean(cell.value) for cell in ws[1]]
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        p = {h: clean(v) for h, v in zip(headers, row)}
        projects.append(p)
    return projects

def load_graph(project_number):
    path = os.path.join(GRAPHS_DIR, f"project_{project_number}.json")
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def available_graph_numbers():
    if not os.path.isdir(GRAPHS_DIR):
        return set()
    nums = set()
    for fname in os.listdir(GRAPHS_DIR):
        if fname.startswith("project_") and fname.endswith(".json"):
            try:
                nums.add(int(fname.replace("project_", "").replace(".json", "")))
            except ValueError:
                pass
    return nums

# ─── Auth routes ─────────────────────────────────────────

@app.route("/api/signup", methods=["POST"])
def signup():
    data     = request.get_json()
    email    = (data.get("email") or "").strip().lower()
    name     = (data.get("name") or "").strip()
    phone    = (data.get("phone") or "").strip()
    password = data.get("password") or ""

    if not email or not name or not phone or not password:
        return jsonify({"ok": False, "error": "All fields are required."}), 400

    try:
        supabase.auth.sign_up({
            "email": email,
            "password": password,
            "options": {
                "data": {
                    "full_name": name,
                    "phone":     phone
                }
            }
        })
        return jsonify({
            "ok": True,
            "pending_verification": True,
            "message": "Account created! Please check your email and click the verification link before signing in."
        })
    except Exception as e:
        err = str(e)
        if "already registered" in err.lower() or "already exists" in err.lower():
            return jsonify({"ok": False, "error": "An account with this email already exists."}), 409
        return jsonify({"ok": False, "error": "Signup failed. Please try again."}), 500


@app.route("/api/login", methods=["POST"])
def login():
    data     = request.get_json()
    email    = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""

    if not email or not password:
        return jsonify({"ok": False, "error": "Email and password are required."}), 400

    try:
        res  = supabase.auth.sign_in_with_password({"email": email, "password": password})
        user = res.user

        if user is None:
            return jsonify({"ok": False, "error": "Invalid email or password."}), 401

        if not user.email_confirmed_at:
            return jsonify({
                "ok": False,
                "error": "Your email is not verified yet. Please check your inbox and click the verification link."
            }), 403

        meta = user.user_metadata or {}
        name = meta.get("full_name", email)
        session["logged_in"]     = True
        session["user_id"]       = user.id
        session["email"]         = user.email
        session["name"]          = name
        session["phone"]         = meta.get("phone", "")
        session["auth_provider"] = "email"
        return jsonify({"ok": True, "name": name})

    except Exception as e:
        err = str(e).lower()
        if "invalid" in err or "credentials" in err or "password" in err:
            return jsonify({"ok": False, "error": "Invalid email or password."}), 401
        if "email not confirmed" in err:
            return jsonify({
                "ok": False,
                "error": "Your email is not verified yet. Please check your inbox and click the verification link."
            }), 403
        return jsonify({"ok": False, "error": "Login failed. Please try again."}), 500


@app.route("/api/google_callback", methods=["POST"])
def google_callback():
    data          = request.get_json()
    access_token  = data.get("access_token", "")
    refresh_token = data.get("refresh_token", "")

    if not access_token or not refresh_token:
        return jsonify({"ok": False, "error": "Missing tokens."}), 400

    try:
        res  = supabase.auth.set_session(access_token, refresh_token)
        user = res.user
        if user is None:
            return jsonify({"ok": False, "error": "Invalid Google session."}), 401

        meta     = user.user_metadata or {}
        name     = meta.get("full_name") or meta.get("name") or user.email
        phone    = meta.get("phone", "")
        provider = (user.app_metadata or {}).get("provider", "google")

        session["logged_in"]     = True
        session["user_id"]       = user.id
        session["email"]         = user.email
        session["name"]          = name
        session["phone"]         = phone
        session["auth_provider"] = provider
        return jsonify({"ok": True, "name": name})
    except Exception as e:
        err = str(e)
        return jsonify({"ok": False, "error": f"Google sign-in failed: {err}"}), 500


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/me")
def me():
    return jsonify({
        "logged_in":     session.get("logged_in", False),
        "name":          session.get("name", ""),
        "email":         session.get("email", ""),
        "phone":         session.get("phone", ""),
        "auth_provider": session.get("auth_provider", "email"),
    })


@app.route("/api/profile", methods=["GET"])
@login_required
def get_profile():
    return jsonify({
        "name":          session.get("name", ""),
        "email":         session.get("email", ""),
        "phone":         session.get("phone", ""),
        "auth_provider": session.get("auth_provider", "email"),
    })


@app.route("/api/profile", methods=["POST"])
@login_required
def update_profile():
    data  = request.get_json()
    name  = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()

    if not name:
        return jsonify({"ok": False, "error": "Name cannot be empty."}), 400
    if phone and (not phone.isdigit() or len(phone) != 10):
        return jsonify({"ok": False, "error": "Phone must be exactly 10 digits."}), 400

    try:
        user_id = session["user_id"]
        supabase_admin.auth.admin.update_user_by_id(user_id, {
            "user_metadata": {
                "full_name": name,
                "phone":     phone
            }
        })
        session["name"]  = name
        session["phone"] = phone
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": "Could not update profile. Please try again."}), 500


# ─── Wallet routes ─────────────────────────────────────────

@app.route("/api/wallet")
@login_required
def wallet():
    user_id = session.get("user_id", "")
    return jsonify(get_wallet(user_id))


@app.route("/api/plans")
def plans():
    """Return available plans."""
    return jsonify({"plans": PLANS, "token_cost": PROJECT_TOKEN_COST})


@app.route("/api/topup", methods=["POST"])
@login_required
def topup():
    """
    Called after a Razorpay payment is confirmed.
    Body: { "plan": "starter" | "enterprise" | "direct" }
    For "starter" / "enterprise": credits plan tokens to wallet.
    For "direct": adds 100 tokens directly to wallet (legacy path).
    NOTE: Single-project direct payment (₹100, no wallet) uses /api/direct_project_unlock instead.
    """
    user_id = session.get("user_id", "")
    data    = request.get_json()
    plan    = data.get("plan", "direct")

    if plan not in ("starter", "enterprise", "direct"):
        return jsonify({"ok": False, "error": "Invalid plan."}), 400

    if plan == "direct":
        # Direct top-up: add 100 tokens (for ₹99 one-time project unlock)
        add_tokens(user_id, PROJECT_TOKEN_COST)
        return jsonify({"ok": True, "tokens_added": PROJECT_TOKEN_COST, "wallet": get_wallet(user_id)})

    cfg     = PLANS[plan]
    tokens  = cfg["tokens"]
    add_tokens(user_id, tokens, plan=plan)
    return jsonify({"ok": True, "tokens_added": tokens, "plan": plan, "wallet": get_wallet(user_id)})


@app.route("/api/unlock_with_tokens", methods=["POST"])
@login_required
def unlock_with_tokens():
    """
    Deduct PROJECT_TOKEN_COST tokens from wallet and unlock the project.
    Body: { "project_index": <int> }
    Returns: { ok, wallet } or { ok: false, insufficient_tokens: true }
    """
    user_id = session.get("user_id", "")
    data    = request.get_json()
    idx     = data.get("project_index")

    if idx is None:
        return jsonify({"ok": False, "error": "Missing project_index"}), 400

    idx = int(idx)

    # Already unlocked — no charge
    if is_unlocked(user_id, idx):
        return jsonify({"ok": True, "already_unlocked": True, "wallet": get_wallet(user_id)})

    # Determine source from active plan
    wallet  = get_wallet(user_id)
    plan    = wallet.get("plan") or {}
    plan_id = plan.get("plan") if plan else None
    source  = plan_id if plan_id in ("starter", "enterprise") else "direct"

    ok = deduct_tokens(user_id, PROJECT_TOKEN_COST)
    if not ok:
        return jsonify({
            "ok":                False,
            "insufficient_tokens": True,
            "tokens_available":  wallet.get("tokens", 0),
            "tokens_required":   PROJECT_TOKEN_COST,
        }), 402

    unlock_project(user_id, idx, source=source)
    return jsonify({"ok": True, "wallet": get_wallet(user_id)})


@app.route("/api/direct_project_unlock", methods=["POST"])
@login_required
def direct_project_unlock():
    """
    Called after a ₹100 Razorpay payment for a single project.
    Does NOT touch the token wallet — unlocks the project as a "direct" unlock (5 days).
    Body: { "project_index": <int> }
    """
    user_id = session.get("user_id", "")
    data    = request.get_json()
    idx     = data.get("project_index")

    if idx is None:
        return jsonify({"ok": False, "error": "Missing project_index"}), 400

    idx = int(idx)

    # Unlock project with "direct" source → 5-day access
    unlock_project(user_id, idx, source="direct")
    return jsonify({"ok": True, "wallet": get_wallet(user_id)})


# ─── Legacy manual unlock (kept for admin use) ─────────────

@app.route("/api/unlock_manual", methods=["POST"])
@login_required
def unlock_manual():
    user_id = session.get("user_id", "")
    data    = request.get_json()
    idx     = data.get("project_index")
    if idx is None:
        return jsonify({"ok": False, "error": "Missing project_index"}), 400
    unlock_project(user_id, int(idx), source="direct")
    return jsonify({"ok": True})


# ─── My unlocks ───────────────────────────────────────────

@app.route("/api/my_unlocks")
@login_required
def my_unlocks():
    user_id  = session.get("user_id", "")
    unlocks  = get_unlocked_list(user_id)
    projects = load_projects()
    result   = []
    for u in unlocks:
        idx  = u["project_index"]
        name = projects[idx].get("project_name", f"Project {idx+1}") if idx < len(projects) else f"Project {idx+1}"
        result.append({
            "project_index": idx,
            "project_name":  name,
            "days_left":     u["days_left"],
            "unlocked_at":   u["unlocked_at"],
            "unlock_source": u.get("unlock_source", "direct"),
        })
    return jsonify({"unlocks": result})

# ─── Data routes ─────────────────────────────────────────

@app.route("/api/projects")
@login_required
def api_projects():
    try:
        user_id    = session.get("user_id", "")
        projects   = load_projects()
        graph_nums = available_graph_numbers()
        unlocked   = get_unlocked_set(user_id)
        result = []
        for i, p in enumerate(projects):
            result.append({
                "index":            i,
                "project_name":     p.get("project_name", ""),
                "project_district": p.get("project_district", ""),
                "project_type":     p.get("project_type", ""),
                "has_graph":        (i + 1) in graph_nums,
                "is_locked":        i not in unlocked,
            })
        return jsonify({"projects": result, "total": len(result)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/project_free/<int:index>")
@login_required
def api_project_free(index):
    try:
        projects = load_projects()
        if index < 0 or index >= len(projects):
            return jsonify({"error": "Project not found"}), 404
        return jsonify({"project": projects[index]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/project/<int:index>")
@login_required
def api_project(index):
    user_id = session.get("user_id", "")
    if not is_unlocked(user_id, index):
        return jsonify({"error": "Payment required", "locked": True}), 402
    try:
        projects = load_projects()
        if index < 0 or index >= len(projects):
            return jsonify({"error": "Project not found"}), 404
        return jsonify({"project": projects[index], "graph": load_graph(index + 1)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/")
@app.route("/index.html")
def index():
    return send_from_directory("static", "index.html")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="0.0.0.0")
    args = parser.parse_args()
    print(f"\n  Property Index Explorer → http://{args.host}:{PORT}")
    print(f"  Auth: Supabase ({SUPABASE_URL})\n")
    app.run(host=args.host, debug=False, port=PORT)